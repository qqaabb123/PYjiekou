# -*- coding: utf-8 -*-
"""
@Time ： 2024/4/11 19:55
@Auth ： 荣
@File ：handle_excel.py
@IDE ：PyCharm
@Motto：ABC(Always Be Coding)
"""
from openpyxl import load_workbook
from Common.handle_path import datas_dir

# class HandleExcel:
#     def __init__(self,file_path,sheetname):
#         self.wb = load_workbook(file_path)
#         self.sh = self.wb[sheetname]
#
#
#     def __read_titles(self):
#         titles = []
#         for item in list(self.sh.rows)[0]:    #遍历第一行当中的列
#             titles.append(item.value)
#         return titles
#     def read_all_datas(self):
#         all_datas = []
#         titles = self.__read_titles()
#         for item in list(self.sh.rows)[1:]:  #遍历数据行
#             values = []
#             for val in item:  #获取每一行的值
#                 values.append(val.value)
#             res = dict(zip(titles,values))   #title和每一行数据，打包成res
#             # res["check"] = eval(res["check"])  #将check的字符串转换为字典对象
#             # res["request_data"] = json.load(res["request_data"])  #将请求数据从json字符串转换成字典格式
#
#             all_datas.append(res)
#         return all_datas
#
#     def close_file(self):
#         self.wb.close()
#
#
#
# if __name__ == '__main__':
#     import os
#     from day1.Common.handle_path import datas_dir
#     file_path = os.path.join(datas_dir,"api_cases.xlsx")
#     # print(file_path)
#     exc = HandleExcel(file_path,"login")
#     # print(exc)
#     cases = exc.read_all_datas()
#     exc.close_file()
#     for case in cases:
#         print(case)


class HandleExcel:

    def __init__(self,file_path,sheet_name):
        self.wb = load_workbook(file_path)
        self.sh = self.wb[sheet_name]

    def __read_titles(self):
        titles = []
        for item in list(self.sh.rows)[0]:  # 遍历第1行当中每一列
            titles.append(item.value)
        return titles

    def read_all_datas(self):
        all_datas = []
        titles = self.__read_titles()
        for item in list(self.sh.rows)[1:]:  # 遍历数据行
            values = []
            for val in item:  # 获取每一行的值
                values.append(val.value)
            res = dict(zip(titles, values))  # title和每一行数据，打包成字典
            all_datas.append(res)
        return all_datas


    def close_file(self):
        self.wb.close()


if __name__ == '__main__':
    import os
    file_path = os.path.join(datas_dir, "api_cases.xlsx")
    # print(file_path)
    exc = HandleExcel(file_path,"login")
    cases = exc.read_all_datas()
    exc.close_file()
    for case in cases:
        print(case)